from __future__ import annotations

import csv
from dataclasses import dataclass
from io import BytesIO, StringIO
from pathlib import Path

from openpyxl import load_workbook


MAX_FILE_SIZE_BYTES = 10 * 1024 * 1024
MAX_DATA_ROWS = 10000


@dataclass(slots=True)
class BroadcastFileParserError(Exception):
    message: str

    def __str__(self) -> str:
        return self.message


async def parse_import_file(file_name: str, payload: bytes) -> dict[str, object]:
    if not payload:
        raise BroadcastFileParserError('导入文件为空，请检查文件内容')

    if len(payload) > MAX_FILE_SIZE_BYTES:
        raise BroadcastFileParserError('文件大小超过限制，请上传 10MB 以内的文件')

    suffix = Path(file_name).suffix.lower()
    if suffix == '.csv':
        return _parse_csv(payload)
    if suffix == '.xlsx':
        return _parse_xlsx(payload)
    raise BroadcastFileParserError('不支持的文件格式，请上传 CSV 或 XLSX 文件')


def _parse_csv(payload: bytes) -> dict[str, object]:
    text = payload.decode('utf-8-sig')
    reader = csv.reader(StringIO(text))
    return _parse_table('csv', None, list(reader))


def _parse_xlsx(payload: bytes) -> dict[str, object]:
    try:
        workbook = load_workbook(
            filename=BytesIO(payload),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        raise BroadcastFileParserError('导入文件已损坏，无法读取，请重新导出后再试') from exc

    try:
        worksheet = workbook.worksheets[0] if workbook.worksheets else None
        if worksheet is None:
            raise BroadcastFileParserError('导入文件没有可读取的数据')

        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([_stringify_cell(value) for value in row])
        return _parse_table('xlsx', worksheet.title, rows)
    finally:
        workbook.close()


def _parse_table(
    file_type: str,
    worksheet_name: str | None,
    rows: list[list[str]],
) -> dict[str, object]:
    if not rows:
        raise BroadcastFileParserError('导入文件没有可读取的数据')

    headers = [_normalize_header(value) for value in rows[0]]
    _validate_headers(headers)

    parsed_rows: list[dict[str, object]] = []
    for row_index, row in enumerate(rows[1:], start=2):
        values = [_normalize_cell(value) for value in row]
        if len(values) < len(headers):
            values.extend([''] * (len(headers) - len(values)))
        if len(values) > len(headers):
            values = values[: len(headers)]
        if all(value == '' for value in values):
            continue
        parsed_rows.append(
            {
                'source_row_number': row_index,
                'raw_data': dict(zip(headers, values, strict=False)),
            }
        )

    if len(parsed_rows) > MAX_DATA_ROWS:
        raise BroadcastFileParserError('导入数据超过 10000 行上限，请拆分后重试')

    return {
        'file_type': file_type,
        'worksheet_name': worksheet_name,
        'headers': headers,
        'rows': parsed_rows,
    }


def _validate_headers(headers: list[str]) -> None:
    seen: set[str] = set()
    for header in headers:
        if not header:
            raise BroadcastFileParserError('导入文件存在空字段名，请检查表头后重试')
        if header in seen:
            raise BroadcastFileParserError(f'导入文件存在重复字段：{header}')
        seen.add(header)


def _normalize_header(value: str) -> str:
    return _normalize_cell(value).lstrip('\ufeff')


def _normalize_cell(value: object) -> str:
    if value is None:
        return ''
    return str(value).strip()


def _stringify_cell(value: object) -> str:
    if value is None:
        return ''
    return str(value)
